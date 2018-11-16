from openpyxl import *
import baostock as bs
import pandas as pd
import numpy as np
from datetime import *
from time import mktime, strptime
from math import log2

# 区间列表
section_list = [[0, 10], [5, 6], [5, 7], [5, 8], [5, 9], [5, 10]]
# 股票代码列表
code_list = ["600108", "600896", "600897", "600642", "600719", "600769", "600863", "600649", "600749", "600692",
             "600098", "600644", "600726", "600780", "600864", "600054", "600798", "600011", "600116", "600674",
             "600744", "600795", "600886", "600138", "600068", "600758", "000002", "000031", "600684", "600743",
             "600766", "600807", "600895", "600419", "600489", "600610", "600750", "600810", "600819", "600830",
             "600889", "600321", "600051", "600531", ]
std_code_list = ["000001", "399001"]
# 000001上证指数 399001深证指数
# 恐怖事件日期
terror_date_list = ['2014-06-15', '2014-08-09', '2017-10-14', '2016-12-10', '2014-08-20', '2014-07-17', '2009-01-17',
                    '2014-08-15', '2016-07-02', '2017-11-24']
# 自然灾害日期
date_list = ['2009-06-29', '2011-07-23', '2014-08-02', '2015-08-12']
# 有具体数据的股票数
NUM_CODE_DETAILS = 10


# 初始化股票数据
def get_stock_data():
    # 登陆系统
    lg = bs.login()
    # 显示登陆返回信息
    print('login respond error_code:' + lg.error_code)
    print('login respond  error_msg:' + lg.error_msg)

    # 获取沪深A股历史K线数据
    # date	交易所行情日期
    # code	证券代码
    # open	开盘价
    # high	最高价
    # low	最低价
    # close	收盘价
    # preclose	昨日收盘价
    for code in code_list + std_code_list:
        if code is '399001':
            code = "sz." + code
        else:
            code = "sh." + code
        # 自然灾害日期
        # 2009-06-29
        rs = bs.query_history_k_data(code, "date,code,open,high,low,close,preclose", start_date='2009-06-19',
                                     end_date='2009-07-10')
        print('query_history_k_data respond error_code:' + rs.error_code)
        print('query_history_k_data respond error_msg:' + rs.error_msg)
        data_list = []
        while (rs.error_code == '0') & rs.next():
            data_list.append(rs.get_row_data())
        result = pd.DataFrame(data_list, columns=rs.fields)
        result.to_excel(r'./stock_data/' + str(code) + '_' + str(date_list[0]) + '.xlsx', index=False)
        # 2011-07-23
        rs = bs.query_history_k_data(code, "date,code,open,high,low,close,preclose", start_date='2011-07-13',
                                     end_date='2011-08-02')
        print('query_history_k_data respond error_code:' + rs.error_code)
        print('query_history_k_data respond error_msg:' + rs.error_msg)
        data_list = []
        while (rs.error_code == '0') & rs.next():
            # 获取一条记录，将记录合并在一起
            data_list.append(rs.get_row_data())
        result = pd.DataFrame(data_list, columns=rs.fields)
        result.to_excel(r'./stock_data/' + str(code) + '_' + str(date_list[1]) + '.xlsx', index=False)
        # 2014-08-02
        rs = bs.query_history_k_data(code, "date,code,open,high,low,close,preclose", start_date='2014-07-23',
                                     end_date='2014-08-12')
        print('query_history_k_data respond error_code:' + rs.error_code)
        print('query_history_k_data respond error_msg:' + rs.error_msg)
        data_list = []
        while (rs.error_code == '0') & rs.next():
            # 获取一条记录，将记录合并在一起
            data_list.append(rs.get_row_data())
        result = pd.DataFrame(data_list, columns=rs.fields)
        result.to_excel(r'./stock_data/' + str(code) + '_' + str(date_list[2]) + '.xlsx', index=False)
        # 2015-08-12
        rs = bs.query_history_k_data(code, "date,code,open,high,low,close,preclose", start_date='2015-08-02',
                                     end_date='2015-08-22')
        print('query_history_k_data respond error_code:' + rs.error_code)
        print('query_history_k_data respond error_msg:' + rs.error_msg)
        data_list = []
        while (rs.error_code == '0') & rs.next():
            # 获取一条记录，将记录合并在一起
            data_list.append(rs.get_row_data())
        result = pd.DataFrame(data_list, columns=rs.fields)
        result.to_excel(r'./stock_data/' + str(code) + '_' + str(date_list[3]) + '.xlsx', index=False)
        # 恐怖事件日期
        # 2014-06-15
        rs = bs.query_history_k_data(code, "date,code,open,high,low,close,preclose", start_date='2014-06-05',
                                     end_date='2014-06-25')
        print('query_history_k_data respond error_code:' + rs.error_code)
        print('query_history_k_data respond error_msg:' + rs.error_msg)
        data_list = []
        while (rs.error_code == '0') & rs.next():
            # 获取一条记录，将记录合并在一起
            data_list.append(rs.get_row_data())
        result = pd.DataFrame(data_list, columns=rs.fields)
        result.to_excel(r'./stock_data/' + str(code) + '_' + str(terror_date_list[0]) + '.xlsx', index=False)
        # 2014-08-09
        rs = bs.query_history_k_data(code, "date,code,open,high,low,close,preclose", start_date='2014-07-30',
                                     end_date='2014-08-19')
        print('query_history_k_data respond error_code:' + rs.error_code)
        print('query_history_k_data respond error_msg:' + rs.error_msg)
        data_list = []
        while (rs.error_code == '0') & rs.next():
            # 获取一条记录，将记录合并在一起
            data_list.append(rs.get_row_data())
        result = pd.DataFrame(data_list, columns=rs.fields)
        result.to_excel(r'./stock_data/' + str(code) + '_' + str(terror_date_list[1]) + '.xlsx', index=False)
        # 2017-10-14
        rs = bs.query_history_k_data(code, "date,code,open,high,low,close,preclose", start_date='2017-10-04',
                                     end_date='2017-10-24')
        print('query_history_k_data respond error_code:' + rs.error_code)
        print('query_history_k_data respond error_msg:' + rs.error_msg)
        data_list = []
        while (rs.error_code == '0') & rs.next():
            # 获取一条记录，将记录合并在一起
            data_list.append(rs.get_row_data())
        result = pd.DataFrame(data_list, columns=rs.fields)
        result.to_excel(r'./stock_data/' + str(code) + '_' + str(terror_date_list[2]) + '.xlsx', index=False)
        # 2016-12-10
        rs = bs.query_history_k_data(code, "date,code,open,high,low,close,preclose", start_date='2016-11-30',
                                     end_date='2016-12-20')
        print('query_history_k_data respond error_code:' + rs.error_code)
        print('query_history_k_data respond error_msg:' + rs.error_msg)
        data_list = []
        while (rs.error_code == '0') & rs.next():
            # 获取一条记录，将记录合并在一起
            data_list.append(rs.get_row_data())
        result = pd.DataFrame(data_list, columns=rs.fields)
        result.to_excel(r'./stock_data/' + str(code) + '_' + str(terror_date_list[3]) + '.xlsx', index=False)
        # 2014-08-20
        rs = bs.query_history_k_data(code, "date,code,open,high,low,close,preclose", start_date='2014-08-10',
                                     end_date='2014-08-30')
        print('query_history_k_data respond error_code:' + rs.error_code)
        print('query_history_k_data respond error_msg:' + rs.error_msg)
        data_list = []
        while (rs.error_code == '0') & rs.next():
            # 获取一条记录，将记录合并在一起
            data_list.append(rs.get_row_data())
        result = pd.DataFrame(data_list, columns=rs.fields)
        result.to_excel(r'./stock_data/' + str(code) + '_' + str(terror_date_list[4]) + '.xlsx', index=False)
        # 2014-07-17
        rs = bs.query_history_k_data(code, "date,code,open,high,low,close,preclose", start_date='2014-07-07',
                                     end_date='2014-07-27')
        print('query_history_k_data respond error_code:' + rs.error_code)
        print('query_history_k_data respond error_msg:' + rs.error_msg)
        data_list = []
        while (rs.error_code == '0') & rs.next():
            # 获取一条记录，将记录合并在一起
            data_list.append(rs.get_row_data())
        result = pd.DataFrame(data_list, columns=rs.fields)
        result.to_excel(r'./stock_data/' + str(code) + '_' + str(terror_date_list[5]) + '.xlsx', index=False)
        # 2009-01-17
        rs = bs.query_history_k_data(code, "date,code,open,high,low,close,preclose", start_date='2009-01-07',
                                     end_date='2009-01-27')
        print('query_history_k_data respond error_code:' + rs.error_code)
        print('query_history_k_data respond error_msg:' + rs.error_msg)
        data_list = []
        while (rs.error_code == '0') & rs.next():
            # 获取一条记录，将记录合并在一起
            data_list.append(rs.get_row_data())
        result = pd.DataFrame(data_list, columns=rs.fields)
        result.to_excel(r'./stock_data/' + str(code) + '_' + str(terror_date_list[6]) + '.xlsx', index=False)
        # 2014-08-15
        rs = bs.query_history_k_data(code, "date,code,open,high,low,close,preclose", start_date='2014-08-05',
                                     end_date='2014-08-25')
        print('query_history_k_data respond error_code:' + rs.error_code)
        print('query_history_k_data respond error_msg:' + rs.error_msg)
        data_list = []
        while (rs.error_code == '0') & rs.next():
            # 获取一条记录，将记录合并在一起
            data_list.append(rs.get_row_data())
        result = pd.DataFrame(data_list, columns=rs.fields)
        result.to_excel(r'./stock_data/' + str(code) + '_' + str(terror_date_list[7]) + '.xlsx', index=False)
        # 2016-07-02
        rs = bs.query_history_k_data(code, "date,code,open,high,low,close,preclose", start_date='2016-06-22',
                                     end_date='2016-07-12')
        print('query_history_k_data respond error_code:' + rs.error_code)
        print('query_history_k_data respond error_msg:' + rs.error_msg)
        data_list = []
        while (rs.error_code == '0') & rs.next():
            # 获取一条记录，将记录合并在一起
            data_list.append(rs.get_row_data())
        result = pd.DataFrame(data_list, columns=rs.fields)
        result.to_excel(r'./stock_data/' + str(code) + '_' + str(terror_date_list[8]) + '.xlsx', index=False)
        # 2017-11-24
        rs = bs.query_history_k_data(code, "date,code,open,high,low,close,preclose", start_date='2017-11-14',
                                     end_date='2017-12-04')
        print('query_history_k_data respond error_code:' + rs.error_code)
        print('query_history_k_data respond error_msg:' + rs.error_msg)
        data_list = []
        while (rs.error_code == '0') & rs.next():
            # 获取一条记录，将记录合并在一起
            data_list.append(rs.get_row_data())
        result = pd.DataFrame(data_list, columns=rs.fields)
        result.to_excel(r'./stock_data/' + str(code) + '_' + str(terror_date_list[9]) + '.xlsx', index=False)
    # 登出系统
    bs.logout()


# 计算正常收益模型的股票/指数收益率 区间[-5,5] 共11天
def cal_RK(code, cur_date):
    cur_path = stock_path(code, cur_date)
    cur_book = load_workbook(cur_path)
    cur_sheet = cur_book.worksheets[0]
    cur_nrow = cur_sheet.max_row
    cur_ncol = cur_sheet.max_column
    cur_first_row_list = []
    for i in range(1, cur_ncol + 1):
        cur_first_row_list.append(cur_sheet.cell(1, i).value)
    ncol_date = cur_first_row_list.index("date") + 1
    ncol_close = cur_first_row_list.index("close") + 1
    ncol_preclose = cur_first_row_list.index("preclose") + 1
    cur_date = strptime(str(cur_date), '%Y-%m-%d')
    value_list1 = []
    value_list2 = []
    for i in range(2, cur_nrow + 1):
        stock_date = strptime(str(cur_sheet.cell(i, ncol_date).value), '%Y-%m-%d')
        if int(mktime(cur_date) > mktime(stock_date)):
            value = (float(cur_sheet.cell(i, ncol_close).value) - float(
                cur_sheet.cell(i, ncol_preclose).value)) / float(cur_sheet.cell(i, ncol_preclose).value)
            value_list1.append(value)
        if int(mktime(cur_date) <= mktime(stock_date)):
            value = (float(cur_sheet.cell(i, ncol_close).value) - float(
                cur_sheet.cell(i, ncol_preclose).value)) / float(cur_sheet.cell(i, ncol_preclose).value)
            value_list2.append(value)
    len_list = len(value_list1)
    result_list1 = []
    result_list2 = []
    # 选取中间的11天
    for i in range(len_list - 5, len_list):
        result_list1.append(value_list1[i])
    for i in range(6):
        result_list2.append(value_list2[i])
    return result_list1 + result_list2


# 计算每个股票每个日期的日收益率 区间[-5,5] 共11天
def get_matrix_RK():
    code_matrix = {}  # 指定股票的日收益率
    std_matrix = {}  # 市场的指数收益率
    for code in code_list:
        for cur_date in date_list:
            result_list = cal_RK(code, cur_date)
            for i in range(len(result_list)):
                code_matrix[code, cur_date, -5 + i] = result_list[i]
    for std_type in range(2):
        for cur_date in date_list:
            result_list = cal_RK(std_code_list[std_type], cur_date)
            for i in range(len(result_list)):
                std_matrix[std_type, cur_date, -5 + i] = result_list[i]
    return [code_matrix, std_matrix]


# 矩阵相减计算超额收益率
def cal_Ex(m1, m2):
    exception_matrix = {}
    for code in code_list:
        for cur_date in date_list:
            if code[0] is '6':
                std_type = 0
            elif code[0] is '0':
                std_type = 1
            else:
                std_type = 2
            for i in range(-5, 6):
                exception_matrix[code, cur_date, i] = m1[code, cur_date, i] - m2[std_type, cur_date, i]
    return exception_matrix


# 将同一时间的所有股票异常收益率累加求平均Ex的平均
def cal_AR(ex):
    AR_matrix = {}
    for cur_date in date_list:
        for i in range(-5, 6):
            total = 0
            for code in code_list:
                total += ex[code, cur_date, i]
            AR_matrix[cur_date, i] = total / len(code_list)
    return AR_matrix


# 计算平均异常收益率
def cal_AARt(ar_matrix, cur_date, start, end):
    total = 0
    for i in range(start, end + 1):
        total += ar_matrix[cur_date, i]
    return total / (end - start + 1)


# 计算累积异常收益率
def cal_CARt(cur_date, start, end):
    total = 0
    for code in code_list:
        if code[0] is '6':
            std_code = '000001'
        else:
            std_code = '399001'
        cur_path = stock_path(code, cur_date)
        cur_book = load_workbook(cur_path)
        cur_sheet = cur_book.worksheets[0]
        cur_nrow = cur_sheet.max_row
        cur_ncol = cur_sheet.max_column
        cur_first_row_list = []
        for i in range(1, cur_ncol + 1):
            cur_first_row_list.append(cur_sheet.cell(1, i).value)
        ncol_date = cur_first_row_list.index("date") + 1
        ncol_close = cur_first_row_list.index("close") + 1
        ncol_preclose = cur_first_row_list.index("preclose") + 1
        _cur_date = strptime(str(cur_date), '%Y-%m-%d')
        value_list1 = []
        value_list2 = []
        for i in range(2, cur_nrow + 1):
            stock_date = strptime(str(cur_sheet.cell(i, ncol_date).value), '%Y-%m-%d')
            if int(mktime(_cur_date) > mktime(stock_date)):
                value_list1.append(
                    [float(cur_sheet.cell(i, ncol_preclose).value), float(cur_sheet.cell(i, ncol_close).value)])
            if int(mktime(_cur_date) <= mktime(stock_date)):
                value_list2.append(
                    [float(cur_sheet.cell(i, ncol_preclose).value), float(cur_sheet.cell(i, ncol_close).value)])
        len_list = len(value_list1)
        result_list1 = []
        result_list2 = []
        # 选取中间的11天
        for i in range(len_list - 5, len_list):
            result_list1.append(value_list1[i])
        for i in range(6):
            result_list2.append(value_list2[i])
        final_code_list = result_list1 + result_list2
        code_pre_close = final_code_list[start + 5][0]
        code_close = final_code_list[end + 5][1]
        CR = code_close / code_pre_close - 1

        cur_path = stock_path(std_code, cur_date)
        cur_book = load_workbook(cur_path)
        cur_sheet = cur_book.worksheets[0]
        cur_nrow = cur_sheet.max_row
        cur_ncol = cur_sheet.max_column
        cur_first_row_list = []
        for i in range(1, cur_ncol + 1):
            cur_first_row_list.append(cur_sheet.cell(1, i).value)
        ncol_date = cur_first_row_list.index("date") + 1
        ncol_close = cur_first_row_list.index("close") + 1
        ncol_preclose = cur_first_row_list.index("preclose") + 1
        _cur_date = strptime(str(cur_date), '%Y-%m-%d')
        value_list1 = []
        value_list2 = []
        for i in range(2, cur_nrow + 1):
            stock_date = strptime(str(cur_sheet.cell(i, ncol_date).value), '%Y-%m-%d')
            if int(mktime(_cur_date) > mktime(stock_date)):
                value_list1.append(
                    [float(cur_sheet.cell(i, ncol_preclose).value), float(cur_sheet.cell(i, ncol_close).value)])
            if int(mktime(_cur_date) <= mktime(stock_date)):
                value_list2.append(
                    [float(cur_sheet.cell(i, ncol_preclose).value), float(cur_sheet.cell(i, ncol_close).value)])
        len_list = len(value_list1)
        result_list1 = []
        result_list2 = []
        # 选取中间的11天
        for i in range(len_list - 5, len_list):
            result_list1.append(value_list1[i])
        for i in range(6):
            result_list2.append(value_list2[i])
        final_code_list = result_list1 + result_list2
        code_pre_close = final_code_list[start + 5][0]
        code_close = final_code_list[end + 5][1]
        std_CR = code_close / code_pre_close - 1
        total += CR - std_CR
    return total / len(code_list)


# 计算日振幅平均
def cal_DA(code, cur_date):
    cur_path = stock_path(code, cur_date)
    try:
        cur_book = load_workbook(cur_path)
    except Exception:
        return 0
    else:
        cur_sheet = cur_book.worksheets[0]
        cur_nrow = cur_sheet.max_row
        cur_ncol = cur_sheet.max_column
        cur_first_row_list = []
        for i in range(1, cur_ncol + 1):
            cur_first_row_list.append(cur_sheet.cell(1, i).value)
        ncol_date = cur_first_row_list.index("date") + 1
        ncol_high = cur_first_row_list.index("high") + 1
        ncol_low = cur_first_row_list.index("low") + 1
        ncol_preclose = cur_first_row_list.index("preclose") + 1
        cur_date = strptime(str(cur_date), '%Y-%m-%d')
        value_list1 = []
        value_list2 = []
        for i in range(2, cur_nrow + 1):
            stock_date = strptime(str(cur_sheet.cell(i, ncol_date).value), '%Y-%m-%d')
            # 事前
            if int(mktime(cur_date) > mktime(stock_date)):
                value = (float(cur_sheet.cell(i, ncol_high).value) - float(
                    cur_sheet.cell(i, ncol_low).value)) / float(cur_sheet.cell(i, ncol_preclose).value)
                value_list1.append(value)
            # 事后
            if int(mktime(cur_date) <= mktime(stock_date)):
                value = (float(cur_sheet.cell(i, ncol_high).value) - float(
                    cur_sheet.cell(i, ncol_low).value)) / float(cur_sheet.cell(i, ncol_preclose).value)
                value_list2.append(value)
        len_list1 = len(value_list1)
        result_list1 = []
        result_list2 = []
        for i in range(len_list1 - 5, len_list1):
            result_list1.append(value_list1[i])
        for j in range(5):
            result_list2.append(value_list2[j])
        return [result_list1, result_list2]


def cal_ADA():
    ADA_matrix = {}
    for code in code_list:
        for cur_date in date_list:
            result = cal_DA(code, cur_date)
            list1 = result[0]  # (-5, -1)
            list2 = result[1]  # (0, 4)
            ave1 = 1 / 5 * sum(list1)
            ave2 = 1 / 5 * sum(list2)
            ADA_matrix[code, cur_date, 0] = ave1
            ADA_matrix[code, cur_date, 1] = ave2
    return ADA_matrix


# 计算日振幅均值和标准差
def ADA_ave_std(ada):
    ADA_as_matrix = {}
    for i in range(2):
        for cur_date in date_list:
            data_list = []
            for code in code_list:
                data_list.append(ada[code, cur_date, i])
            ave = sum(data_list) / len(data_list)
            std = np.std(data_list, ddof=1)
            ADA_as_matrix[cur_date, i] = [ave, std]
    return ADA_as_matrix


# 计算短期波动
def cal_VOL(code, cur_date):
    cur_path = stock_path(code, cur_date)
    try:
        cur_book = load_workbook(cur_path)
    except Exception:
        return 0
    else:
        cur_sheet = cur_book.worksheets[0]
        cur_nrow = cur_sheet.max_row
        cur_ncol = cur_sheet.max_column
        cur_first_row_list = []
        for i in range(1, cur_ncol + 1):
            cur_first_row_list.append(cur_sheet.cell(1, i).value)
        ncol_date = cur_first_row_list.index("date") + 1
        ncol_high = cur_first_row_list.index("high") + 1
        ncol_low = cur_first_row_list.index("low") + 1
        cur_date = strptime(str(cur_date), '%Y-%m-%d')
        value_list1 = []
        value_list2 = []
        for i in range(2, cur_nrow + 1):
            stock_date = strptime(str(cur_sheet.cell(i, ncol_date).value), '%Y-%m-%d')
            # 事前
            if int(mktime(cur_date) > mktime(stock_date)):
                value = (float(cur_sheet.cell(i, ncol_high).value) - float(
                    cur_sheet.cell(i, ncol_low).value)) / float(cur_sheet.cell(i, ncol_low).value)
                value_list1.append(value)
            # 事后
            if int(mktime(cur_date) <= mktime(stock_date)):
                value = (float(cur_sheet.cell(i, ncol_high).value) - float(
                    cur_sheet.cell(i, ncol_low).value)) / float(cur_sheet.cell(i, ncol_low).value)
                value_list2.append(value)
        len_list1 = len(value_list1)
        result_list1 = []
        result_list2 = []
        for i in range(len_list1 - 5, len_list1):
            result_list1.append(value_list1[i])
        for j in range(5):
            result_list2.append(value_list2[j])
        return [result_list1, result_list2]


def cal_AVOL():
    VOL_matrix = {}
    for code in code_list:
        for cur_date in date_list:
            result = cal_VOL(code, cur_date)
            list1 = result[0]  # (-5, -1)
            list2 = result[1]  # (0, 4)
            ave1 = 1 / 5 * sum(list1)
            ave2 = 1 / 5 * sum(list2)
            VOL_matrix[code, cur_date, 0] = ave1
            VOL_matrix[code, cur_date, 1] = ave2
    return VOL_matrix


def AVOL_ave_std(vol):
    VOL_as_matrix = {}
    for i in range(2):
        for cur_date in date_list:
            data_list = []
            for code in code_list:
                data_list.append(vol[code, cur_date, i])
            ave = sum(data_list) / len(data_list)
            std = np.std(data_list, ddof=1)
            VOL_as_matrix[cur_date, i] = [ave, std]
    return VOL_as_matrix


# 计算累计日常收益
def CAR_ave_std(cur_date, start, end):
    CAR_list = []
    for code in code_list:
        if code[0] is '6':
            std_code = '000001'
        else:
            std_code = '399001'
        cur_path = stock_path(code, cur_date)
        cur_book = load_workbook(cur_path)
        cur_sheet = cur_book.worksheets[0]
        cur_nrow = cur_sheet.max_row
        cur_ncol = cur_sheet.max_column
        cur_first_row_list = []
        for i in range(1, cur_ncol + 1):
            cur_first_row_list.append(cur_sheet.cell(1, i).value)
        ncol_date = cur_first_row_list.index("date") + 1
        ncol_close = cur_first_row_list.index("close") + 1
        ncol_preclose = cur_first_row_list.index("preclose") + 1
        _cur_date = strptime(str(cur_date), '%Y-%m-%d')
        value_list1 = []
        value_list2 = []
        for i in range(2, cur_nrow + 1):
            stock_date = strptime(str(cur_sheet.cell(i, ncol_date).value), '%Y-%m-%d')
            if int(mktime(_cur_date) > mktime(stock_date)):
                value_list1.append(
                    [float(cur_sheet.cell(i, ncol_preclose).value), float(cur_sheet.cell(i, ncol_close).value)])
            if int(mktime(_cur_date) <= mktime(stock_date)):
                value_list2.append(
                    [float(cur_sheet.cell(i, ncol_preclose).value), float(cur_sheet.cell(i, ncol_close).value)])
        len_list = len(value_list1)
        result_list1 = []
        result_list2 = []
        # 选取中间的11天
        for i in range(len_list - 5, len_list):
            result_list1.append(value_list1[i])
        for i in range(6):
            result_list2.append(value_list2[i])
        final_code_list = result_list1 + result_list2
        code_pre_close = final_code_list[start + 5][0]
        code_close = final_code_list[end + 5][1]
        CR = code_close / code_pre_close - 1

        cur_path = stock_path(std_code, cur_date)
        cur_book = load_workbook(cur_path)
        cur_sheet = cur_book.worksheets[0]
        cur_nrow = cur_sheet.max_row
        cur_ncol = cur_sheet.max_column
        cur_first_row_list = []
        for i in range(1, cur_ncol + 1):
            cur_first_row_list.append(cur_sheet.cell(1, i).value)
        ncol_date = cur_first_row_list.index("date") + 1
        ncol_close = cur_first_row_list.index("close") + 1
        ncol_preclose = cur_first_row_list.index("preclose") + 1
        _cur_date = strptime(str(cur_date), '%Y-%m-%d')
        value_list1 = []
        value_list2 = []
        for i in range(2, cur_nrow + 1):
            stock_date = strptime(str(cur_sheet.cell(i, ncol_date).value), '%Y-%m-%d')
            if int(mktime(_cur_date) > mktime(stock_date)):
                value_list1.append(
                    [float(cur_sheet.cell(i, ncol_preclose).value), float(cur_sheet.cell(i, ncol_close).value)])
            if int(mktime(_cur_date) <= mktime(stock_date)):
                value_list2.append(
                    [float(cur_sheet.cell(i, ncol_preclose).value), float(cur_sheet.cell(i, ncol_close).value)])
        len_list = len(value_list1)
        result_list1 = []
        result_list2 = []
        # 选取中间的11天
        for i in range(len_list - 5, len_list):
            result_list1.append(value_list1[i])
        for i in range(6):
            result_list2.append(value_list2[i])
        final_code_list = result_list1 + result_list2
        code_pre_close = final_code_list[start + 5][0]
        code_close = final_code_list[end + 5][1]
        std_CR = code_close / code_pre_close - 1
        CAR_list.append(CR - std_CR)
    return [sum(CAR_list) / len(code_list), np.std(CAR_list)]


def stock_path(code, cur_date):
    if code is '399001':
        return r'./stock_data/sz.' + code + '_' + cur_date + '.xlsx'
    else:
        return r'./stock_data/sh.' + code + '_' + cur_date + '.xlsx'


# # 获取股票数据
# get_stock_data()

# date: 日期 同时表示事件
# code: 股票代码
# i: 第i期
[code_matrix, std_matrix] = get_matrix_RK()  # 事故发生前后的日收益率
exception_matrix = cal_Ex(code_matrix, std_matrix)  # 计算异常收益率日数据
ar_matrix = cal_AR(exception_matrix)  # 异常收率（日数据按股票累加求平均）
# 输出
for cur_date in date_list:
    print(cur_date + "事件：")
    print("[-5, -1]")
    print("平均异常收益率AAR:" + str(cal_AARt(ar_matrix, cur_date, -5, -1)))
    print("累积日常收益率CAR:" + str(cal_CARt(cur_date, -5, -1)))
    for i in range(5):
        print(i)
        print("平均异常收益率AAR:" + str(cal_AARt(ar_matrix, cur_date, i, i)))
        print("累积日常收益率CAR:" + str(cal_CARt(cur_date, i, i)))
    for j in range(1, 6):
        print("[0, " + str(j) + "]")
        print("平均异常收益率AAR:" + str(cal_AARt(ar_matrix, cur_date, 0, j)))
        print("累积日常收益率CAR:" + str(cal_CARt(cur_date, 0, j)))

ada_matrix = cal_ADA()
ada_as_matrix = ADA_ave_std(ada_matrix)
print("日振幅平均")
for cur_date in date_list:
    print(cur_date + "事件：")
    result_list = ada_as_matrix[cur_date, 0]
    print("时间区间(-5. -1): 均值：" + str(result_list[0]) + " 标准差：" + str(result_list[1]))
    result_list = ada_as_matrix[cur_date, 1]
    print("时间区间(-5. -1): 均值：" + str(result_list[0]) + " 标准差：" + str(result_list[1]))

avol_matrix = cal_AVOL()


avol_as_matrix = AVOL_ave_std(avol_matrix)
print("短期波动")
for cur_date in date_list:
    print(cur_date + "事件：")
    result_list = avol_as_matrix[cur_date, 0]
    print("时间区间(-5. -1): 均值：" + str(result_list[0]) + " 标准差：" + str(result_list[1]))
    result_list = avol_as_matrix[cur_date, 1]
    print("时间区间(-5. -1): 均值：" + str(result_list[0]) + " 标准差：" + str(result_list[1]))

print("日振幅平均")
for cur_date in date_list:
    print(cur_date + "事件：")
    print("时间区间(-5. -1): 均值：" + str(CAR_ave_std(cur_date, -5, -1)[0]) + " 标准差：" + str(CAR_ave_std(cur_date, -5, -1)[1]))
    print("时间区间(0. 4): 均值：" + str(CAR_ave_std(cur_date, 0, 4)[0]) + " 标准差：" + str(CAR_ave_std(cur_date, 0, 4)[1]))


# # 获取回归分析中的ada(i)
# def get_ada_i(ada_matrix):
#     print("获取回归分析中的ada(i)")
#     ada_i_matrix = {}
#     for code in code_list:
#         for cur_date in date_list:
#             ave1 = ada_matrix[code, cur_date, 0]
#             ave2 = ada_matrix[code, cur_date, 1]
#             ada_i_matrix[code, cur_date] = 1 / 2 * (ave1 + ave2)
#     return ada_i_matrix
#
#
# # 获取回归分析中的vol(i)
# def get_vol_i(avol_matrix):
#     print("获取回归分析中的vol(i)")
#     vol_i_matrix = {}
#     for code in code_list:
#         for cur_date in date_list:
#             ave1 = avol_matrix[code, cur_date, 0]
#             ave2 = avol_matrix[code, cur_date, 1]
#             vol_i_matrix[code, cur_date] = 1 / 2 * (ave1 + ave2)
#     return vol_i_matrix
#
#
# # 获取回归分析中的car(i)
# def get_car_i():
#     print("获取回归分析中的car(i)")
#     car_i_matrix = {}
#     for code in code_list:
#         for cur_date in date_list:
#             if code[0] is '6':
#                 std_code = '000001'
#             else:
#                 std_code = '399001'
#             cur_path = stock_path(code, cur_date)
#             cur_book = load_workbook(cur_path)
#             cur_sheet = cur_book.worksheets[0]
#             cur_nrow = cur_sheet.max_row
#             cur_ncol = cur_sheet.max_column
#             cur_first_row_list = []
#             for i in range(1, cur_ncol + 1):
#                 cur_first_row_list.append(cur_sheet.cell(1, i).value)
#             ncol_date = cur_first_row_list.index("date") + 1
#             ncol_close = cur_first_row_list.index("close") + 1
#             ncol_preclose = cur_first_row_list.index("preclose") + 1
#             _cur_date = strptime(str(cur_date), '%Y-%m-%d')
#             value_list1 = []
#             value_list2 = []
#             for i in range(2, cur_nrow + 1):
#                 stock_date = strptime(str(cur_sheet.cell(i, ncol_date).value), '%Y-%m-%d')
#                 if int(mktime(_cur_date) > mktime(stock_date)):
#                     value_list1.append(
#                         [float(cur_sheet.cell(i, ncol_preclose).value), float(cur_sheet.cell(i, ncol_close).value)])
#                 if int(mktime(_cur_date) <= mktime(stock_date)):
#                     value_list2.append(
#                         [float(cur_sheet.cell(i, ncol_preclose).value), float(cur_sheet.cell(i, ncol_close).value)])
#             len_list = len(value_list1)
#             result_list1 = []
#             result_list2 = []
#             # 选取中间的11天
#             for i in range(len_list - 5, len_list):
#                 result_list1.append(value_list1[i])
#             for i in range(6):
#                 result_list2.append(value_list2[i])
#             final_code_list = result_list1 + result_list2
#             code_pre_close = final_code_list[0][0]
#             code_close = final_code_list[10][1]
#             CR = code_close / code_pre_close - 1
#
#             cur_path = stock_path(std_code, cur_date)
#             cur_book = load_workbook(cur_path)
#             cur_sheet = cur_book.worksheets[0]
#             cur_nrow = cur_sheet.max_row
#             cur_ncol = cur_sheet.max_column
#             cur_first_row_list = []
#             for i in range(1, cur_ncol + 1):
#                 cur_first_row_list.append(cur_sheet.cell(1, i).value)
#             ncol_date = cur_first_row_list.index("date") + 1
#             ncol_close = cur_first_row_list.index("close") + 1
#             ncol_preclose = cur_first_row_list.index("preclose") + 1
#             _cur_date = strptime(str(cur_date), '%Y-%m-%d')
#             value_list1 = []
#             value_list2 = []
#             for i in range(2, cur_nrow + 1):
#                 stock_date = strptime(str(cur_sheet.cell(i, ncol_date).value), '%Y-%m-%d')
#                 if int(mktime(_cur_date) > mktime(stock_date)):
#                     value_list1.append(
#                         [float(cur_sheet.cell(i, ncol_preclose).value), float(cur_sheet.cell(i, ncol_close).value)])
#                 if int(mktime(_cur_date) <= mktime(stock_date)):
#                     value_list2.append(
#                         [float(cur_sheet.cell(i, ncol_preclose).value), float(cur_sheet.cell(i, ncol_close).value)])
#             len_list = len(value_list1)
#             result_list1 = []
#             result_list2 = []
#             # 选取中间的11天
#             for i in range(len_list - 5, len_list):
#                 result_list1.append(value_list1[i])
#             for i in range(6):
#                 result_list2.append(value_list2[i])
#             final_code_list = result_list1 + result_list2
#             code_pre_close = final_code_list[0][0]
#             code_close = final_code_list[10][1]
#             std_CR = code_close / code_pre_close - 1
#             result = CR - std_CR
#             car_i_matrix[code, cur_date] = result
#     return car_i_matrix


# # 获取第i只股票的ada vol car
# ada_i_matrix = get_ada_i(ada_matrix)
# vol_i_matrix = get_vol_i(avol_matrix)
# car_i_matrix = get_car_i()
# for i in range(NUM_CODE_DETAILS):
#     for cur_date in date_list:
#         code = code_list[i]
#         print(cur_date + " " + code)
#         print("ADA:" + str(ada_i_matrix[code, cur_date]) + "VOL:" + str(vol_i_matrix[code, cur_date]) + "CAR:" + str(
#             car_i_matrix[code, cur_date]))

# # 在拿到每一只股票的变量之后！我们就可以导出一个Excel！然后SPSS回归分析！！！
# wb = Workbook()
# year_list = ['09', '11', '14', '15']
# ncol_kd = 12
# ncol_wd = 13
# ncol_ed = 14
# ncol_pd = 15
# stock_book = load_workbook(r'./competition topic/stock_details.xlsx')
# stock_sheet = stock_book.worksheets[0]
# ncol_data = stock_sheet.max_column  # 当前工作表的列数
# first_row_list = []
# for i in range(1, ncol_data + 1):
#     first_row_list.append(stock_sheet.cell(1, i).value)
# affair_book = load_workbook(r'./competition topic/事故灾害统计表.xlsx')
# affair_sheet = affair_book.worksheets[0]
# # cal ada
# sheet = wb.create_sheet(u'ADA', 0)
# first_row = ['ADA', 'ONE', 'EPS', 'ROA', 'TRA1', 'TRA5', 'INV', 'SIZE', 'ST', 'prov_d', 'nkill_d', 'nwound_d', 'eco_d']
# for i in range(len(first_row)):
#     sheet.cell(1, i + 1).value = first_row[i]
# cur_row = 2
# for i in range(NUM_CODE_DETAILS):  # i表示第几只股票
#     for j in range(len(date_list)):  # j表示第几件事情
#         print("正在插入ADA第" + str(cur_row) + "行")
#         income = float(stock_sheet.cell(i+2, first_row_list.index('income_' + year_list[j]) + 1).value)
#         total = float(stock_sheet.cell(i+2, first_row_list.index('total_' + year_list[j]) + 1).value)
#         ada = ada_i_matrix[code_list[i], date_list[j]]
#         one = 1
#         eps = stock_sheet.cell(i+2, first_row_list.index('eps_' + year_list[j]) + 1).value
#         roa = income / total
#         tra1 = stock_sheet.cell(i+2, first_row_list.index('tra1_' + year_list[j]) + 1).value
#         tra5 = stock_sheet.cell(i+2, first_row_list.index('tra5_' + year_list[j]) + 1).value
#         inv = 0
#         size = log2(total)
#         st = stock_sheet.cell(i+2, first_row_list.index('st') + 1).value
#         prov_d = affair_sheet.cell(j + 2, ncol_pd).value
#         nkill_d = affair_sheet.cell(j + 2, ncol_kd).value
#         nwound_d = affair_sheet.cell(j + 2, ncol_wd).value
#         eco_d = affair_sheet.cell(j + 2, ncol_ed).value
#         new_line = [ada, 1, eps, roa, tra1, tra5, inv, size, st, prov_d, nkill_d, nwound_d, eco_d]
#         for k in range(len(new_line)):
#             sheet.cell(cur_row, k + 1).value = new_line[k]
#         cur_row += 1
# # cal vol
# sheet = wb.create_sheet(u'VOL', 1)
# first_row = ['VOL', 'ONE', 'EPS', 'ROA', 'TRA1', 'TRA5', 'INV', 'SIZE', 'ST', 'prov_d', 'nkill_d', 'nwound_d', 'eco_d']
# for i in range(len(first_row)):
#     sheet.cell(1, i + 1).value = first_row[i]
# cur_row = 2
# for i in range(NUM_CODE_DETAILS):  # i表示第几只股票
#     for j in range(len(date_list)):  # j表示第几件事情
#         print("正在插入VOL第" + str(cur_row) + "行")
#         income = float(stock_sheet.cell(i+2, first_row_list.index('income_' + year_list[j]) + 1).value)
#         total = float(stock_sheet.cell(i+2, first_row_list.index('total_' + year_list[j]) + 1).value)
#         ada = vol_i_matrix[code_list[i], date_list[j]]
#         one = 1
#         eps = stock_sheet.cell(i+2, first_row_list.index('eps_' + year_list[j]) + 1).value
#         roa = income / total
#         tra1 = stock_sheet.cell(i+2, first_row_list.index('tra1_' + year_list[j]) + 1).value
#         tra5 = stock_sheet.cell(i+2, first_row_list.index('tra5_' + year_list[j]) + 1).value
#         inv = 0
#         size = log2(total)
#         st = stock_sheet.cell(i+2, first_row_list.index('st') + 1).value
#         prov_d = affair_sheet.cell(j + 2, ncol_pd).value
#         nkill_d = affair_sheet.cell(j + 2, ncol_kd).value
#         nwound_d = affair_sheet.cell(j + 2, ncol_wd).value
#         eco_d = affair_sheet.cell(j + 2, ncol_ed).value
#         new_line = [ada, 1, eps, roa, tra1, tra5, inv, size, st, prov_d, nkill_d, nwound_d, eco_d]
#         for k in range(len(new_line)):
#             sheet.cell(cur_row, k + 1).value = new_line[k]
#         cur_row += 1
# # cal car
# sheet = wb.create_sheet(u'CAR', 2)
# first_row = ['CAR', 'ONE', 'EPS', 'ROA', 'TRA1', 'TRA5', 'INV', 'SIZE', 'ST', 'prov_d', 'nkill_d', 'nwound_d', 'eco_d']
# for i in range(len(first_row)):
#     sheet.cell(1, i + 1).value = first_row[i]
# cur_row = 2
# for i in range(NUM_CODE_DETAILS):  # i表示第几只股票
#     for j in range(len(date_list)):  # j表示第几件事情
#         print("正在插入CAR第" + str(cur_row) + "行")
#         income = float(stock_sheet.cell(i+2, first_row_list.index('income_' + year_list[j]) + 1).value)
#         total = float(stock_sheet.cell(i+2, first_row_list.index('total_' + year_list[j]) + 1).value)
#         ada = car_i_matrix[code_list[i], date_list[j]]
#         one = 1
#         eps = stock_sheet.cell(i+2, first_row_list.index('eps_' + year_list[j]) + 1).value
#         roa = income / total
#         tra1 = stock_sheet.cell(i+2, first_row_list.index('tra1_' + year_list[j]) + 1).value
#         tra5 = stock_sheet.cell(i+2, first_row_list.index('tra5_' + year_list[j]) + 1).value
#         inv = 0
#         size = log2(total)
#         st = stock_sheet.cell(i+2, first_row_list.index('st') + 1).value
#         prov_d = affair_sheet.cell(j + 2, ncol_pd).value
#         nkill_d = affair_sheet.cell(j + 2, ncol_kd).value
#         nwound_d = affair_sheet.cell(j + 2, ncol_wd).value
#         eco_d = affair_sheet.cell(j + 2, ncol_ed).value
#         new_line = [ada, 1, eps, roa, tra1, tra5, inv, size, st, prov_d, nkill_d, nwound_d, eco_d]
#         for k in range(len(new_line)):
#             sheet.cell(cur_row, k + 1).value = new_line[k]
#         cur_row += 1
# wb.save(r'./competition topic/待回归分析数据.xlsx')
