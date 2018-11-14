from openpyxl import *
import numpy as np
import numpy.linalg as nlg
import pandas as pd
from math import pow, sqrt

global NUM_TYPE_WEAPON
global NUM_TYPE_ATTACK
global NUM_TYPE_TARGET
NUM_TYPE_WEAPON = 13  # 武器信息种类数量
NUM_TYPE_ATTACK = 9  # 攻击信息种类数量
NUM_TYPE_TARGET = 22  # 攻击目标种类数量


# 用于打开文件，保存文件实例
class Data:

    def __init__(self, file_path):
        self.data = None
        self.sheet = None
        print("正在打开" + file_path)
        self.data = load_workbook(file_path)
        self.reference_data = load_workbook(r'./competition topic/年度地区总人数及GDP情况.xlsx')
        print("打开文件: " + file_path)
        self.sheet = self.data.worksheets[0]
        self.reference_sheet = self.reference_data.worksheets[0]
        self.nrow_data = self.sheet.max_row  # 当前工作表的行数
        self.ncol_data = self.sheet.max_column  # 当前工作表的列数
        self.first_row_list = []
        for i in range(1, self.ncol_data + 1):
            self.first_row_list.append(self.sheet.cell(1, i).value)
        # 对用列序号
        self.ncol_iyear = self.first_row_list.index('iyear') + 1
        self.ncol_region = self.first_row_list.index('region') + 1
        self.ncol_attacktype1 = self.first_row_list.index('attacktype1') + 1
        self.ncol_attacktype2 = self.first_row_list.index('attacktype2') + 1
        self.ncol_attacktype3 = self.first_row_list.index('attacktype3') + 1
        self.ncol_weaptype1 = self.first_row_list.index('weaptype1') + 1
        self.ncol_weaptype2 = self.first_row_list.index('weaptype2') + 1
        self.ncol_weaptype3 = self.first_row_list.index('weaptype3') + 1
        self.ncol_targtype1 = self.first_row_list.index('targtype1') + 1
        self.ncol_targtype2 = self.first_row_list.index('targtype2') + 1
        self.ncol_targtype3 = self.first_row_list.index('targtype3') + 1
        self.ncol_nkill = self.first_row_list.index('nkill') + 1
        self.ncol_nwound = self.first_row_list.index('nwound') + 1
        self.ncol_property = self.first_row_list.index('property') + 1
        self.ncol_propextent = self.first_row_list.index('propextent') + 1

    # 数据填充
    def initial_data(self):
        # 填充经济损失为0情况下的经济损失等级
        for i in range(2, self.nrow_data + 1):
            if self.sheet.cell(i, self.ncol_property).value == 0:
                self.sheet.cell(i, self.ncol_propextent).value = 0
            if self.sheet.cell(i, self.ncol_attacktype2).value is None:
                self.sheet.cell(i, self.ncol_attacktype2).value = 0
            if self.sheet.cell(i, self.ncol_attacktype3).value is None:
                self.sheet.cell(i, self.ncol_attacktype3).value = 0
            if self.sheet.cell(i, self.ncol_weaptype2).value is None:
                self.sheet.cell(i, self.ncol_weaptype2).value = 0
            if self.sheet.cell(i, self.ncol_weaptype3).value is None:
                self.sheet.cell(i, self.ncol_weaptype3).value = 0
            if self.sheet.cell(i, self.ncol_targtype2).value is None:
                self.sheet.cell(i, self.ncol_targtype2).value = 0
            if self.sheet.cell(i, self.ncol_targtype3).value is None:
                self.sheet.cell(i, self.ncol_targtype3).value = 0
        self.data.save(r'./competition topic/proceed.xlsx')
        self.data = load_workbook(r'./competition topic/proceed.xlsx')
        self.sheet = self.data.worksheets[0]
        # 根据欧几里得距离计算相似度，匹配填充nkill,nwound,propextent
        # 数据二维列表，n行12列，计算保存相似度
        # [nrow,nyear,region,attacktype1,2,3,weapontype1,2,3,targtype1,2,3]
        data_list = []
        # 需要计算的行号
        num_need_cal_list = []
        # 从first_sheet中提取符合条件的数据保存至data_matrix中（三列均非空）
        for row in range(2, self.nrow_data + 1):
            if self.sheet.cell(row, self.ncol_nkill).value is not None \
                    and self.sheet.cell(row, self.ncol_nwound).value is not None \
                    and self.sheet.cell(row, self.ncol_property).value != -9:
                # 插入到data_matrix中
                insert_list = self.ret_datalist_cal_simi(row)
                data_list.append(insert_list)
            else:
                # 将序号插入到num_need_cal_list中
                num_need_cal_list.append(row)
        # 开始计算
        for i in num_need_cal_list:
            # 临时列表，用于保存当前处理的数据 nrow:原来xls中是第nrow行数据
            # [nrow,nyear,region,attacktype1,2,3,weapontype1,2,3,targtype1,2,3]
            temp_list = self.ret_datalist_cal_simi(i)
            # 相关度列表
            similarity_list = []
            for j in range(len(data_list)):
                similarity_list.append(self.ret_similarity(temp_list, data_list[j]))
            min_similarity = min(similarity_list)
            min_pos = data_list[similarity_list.index(min_similarity)][0]
            # 若i行数据为空，则填入min_pos行数据
            if self.sheet.cell(i, self.ncol_nkill).value is None:
                kill_data = self.sheet.cell(min_pos, self.ncol_nkill).value
                self.sheet.cell(i, self.ncol_nkill).value = kill_data
            if self.sheet.cell(i, self.ncol_nwound).value is None:
                wound_data = self.sheet.cell(min_pos, self.ncol_nwound).value
                self.sheet.cell(i, self.ncol_nwound).value = wound_data
            if self.sheet.cell(i, self.ncol_propextent).value is None:
                prop_data = self.sheet.cell(min_pos, self.ncol_propextent).value
                self.sheet.cell(i, self.ncol_propextent).value = prop_data
        self.data.save(r'./competition topic/proceed.xlsx')
        self.data = load_workbook(r'./competition topic/proceed.xlsx')
        self.sheet = self.data.worksheets[0]

    # 返回填充数据时需要计算相似度的数据
    def ret_datalist_cal_simi(self, row):
        return [row, self.sheet.cell(row, self.ncol_iyear).value, self.sheet.cell(row, self.ncol_region).value,
                self.sheet.cell(row, self.ncol_attacktype1).value,
                self.sheet.cell(row, self.ncol_attacktype2).value,
                self.sheet.cell(row, self.ncol_attacktype3).value, self.sheet.cell(row, self.ncol_weaptype1).value,
                self.sheet.cell(row, self.ncol_weaptype2).value, self.sheet.cell(row, self.ncol_weaptype3).value,
                self.sheet.cell(row, self.ncol_targtype1).value, self.sheet.cell(row, self.ncol_targtype2).value,
                self.sheet.cell(row, self.ncol_targtype3).value]

    # 获取指定表名的数据列（优先选取第一列）
    def get_data_list(self, col_name):
        ncol = self.first_row_list.index(col_name) + 1
        result_list = []
        for i in range(2, self.nrow_data + 1):
            result_list.append(self.sheet.cell(i, ncol).value)
        return result_list

    # 统计指定攻击信息和武器信息情况下的人员伤亡情况
    # 返回伤亡情况数组: 2 * 武器信息种类 * 攻击信息种类
    def get_casualty_by_attack_and_weapon(self):
        death_list = self.get_data_list('nkill')
        print("death_list length: " + str(len(death_list)))
        wound_list = self.get_data_list('nwound')
        print("wound_list length: " + str(len(wound_list)))
        casualty_array = np.zeros((2, NUM_TYPE_WEAPON + 1, NUM_TYPE_ATTACK + 1))
        for j in range(3):
            attack_list = self.get_data_list('attacktype' + str(j + 1))
            weapon_list = self.get_data_list('weaptype' + str(j + 1))
            for i in range(len(death_list)):
                casualty_array[0, int(weapon_list[i]), int(attack_list[i])] = \
                    casualty_array[0][int(weapon_list[i])][int(attack_list[i])] + int(death_list[i])
                casualty_array[1, int(weapon_list[i]), int(attack_list[i])] = \
                    casualty_array[1][int(weapon_list[i])][int(attack_list[i])] + int(wound_list[i])
            attack_list.clear()
            weapon_list.clear()
        return casualty_array

    # 获取单一攻击信息下的伤亡情况
    def get_casualty_by_attack(self):
        num_peop_matrix = self.get_matrix_peop_by_year_and_region()
        result_dict = {}
        attack_times_dict = {}  # 计算某种攻击类型的次数
        for i in range(1, NUM_TYPE_ATTACK + 1):
            result_dict[i] = [0, 0]
            attack_times_dict[i] = 0
        for i in range(2, self.nrow_data + 1):
            iyear = self.sheet.cell(i, self.ncol_iyear).value
            region = self.sheet.cell(i, self.ncol_region).value
            total = num_peop_matrix[iyear, region]
            nkill = self.sheet.cell(i, self.ncol_nkill).value / total
            nwound = self.sheet.cell(i, self.ncol_nwound).value / total
            attack_type1 = self.sheet.cell(i, self.ncol_attacktype1).value
            attack_type2 = self.sheet.cell(i, self.ncol_attacktype2).value
            attack_type3 = self.sheet.cell(i, self.ncol_attacktype3).value
            if attack_type2 == 0:  # 只有一种攻击的情况
                m = result_dict[attack_type1][0]
                n = result_dict[attack_type1][1]
                result_dict[attack_type1] = [m + nkill, n + nwound]
                attack_times_dict[attack_type1] += 1
            elif attack_type3 == 0:  # 有两种攻击的情况
                m = result_dict[attack_type1][0]
                n = result_dict[attack_type1][1]
                o = result_dict[attack_type2][0]
                p = result_dict[attack_type2][1]
                result_dict[attack_type1] = [m + nkill * 1 / 2, n + nwound * 1 / 2]
                result_dict[attack_type2] = [o + nkill * 1 / 2, p + nwound * 1 / 2]
                attack_times_dict[attack_type1] += 1
                attack_times_dict[attack_type2] += 1
            else:  # 有三种攻击的情况
                m = result_dict[attack_type1][0]
                n = result_dict[attack_type1][1]
                o = result_dict[attack_type2][0]
                p = result_dict[attack_type2][1]
                x = result_dict[attack_type3][0]
                y = result_dict[attack_type3][1]
                result_dict[attack_type1] = [m + nkill * 1 / 3, n + nwound * 1 / 3]
                result_dict[attack_type2] = [o + nkill * 1 / 3, p + nwound * 1 / 3]
                result_dict[attack_type3] = [x + nkill * 1 / 3, y + nwound * 1 / 3]
                attack_times_dict[attack_type1] += 1
                attack_times_dict[attack_type2] += 1
                attack_times_dict[attack_type3] += 1
        # 将攻击总威力除去总次数，得出平均威力
        for i in range(1, NUM_TYPE_ATTACK + 1):
            times = attack_times_dict[i]
            m = result_dict[i][0]
            n = result_dict[i][1]
            if times == 0:
                result_dict[i] = [0, 0]
            else:
                result_dict[i] = [m / times * 10000, n / times * 10000]
        return result_dict

    # 根据伤亡情况对攻击信息进行分类
    def get_score_attack_list_by_casualty(self):
        casualty_dict = self.get_casualty_by_attack()
        score_dict = {}
        for i in range(1, NUM_TYPE_ATTACK + 1):
            score_dict[i] = casualty_dict[i][0] * 0.8 + casualty_dict[i][1] * 0.2
        rank_list = self.sort_by_value(score_dict)
        result_dict = {}
        if len(rank_list) == NUM_TYPE_ATTACK:
            result_dict[rank_list[0]] = 5
            result_dict[rank_list[1]] = 4
            result_dict[rank_list[2]] = 3
            result_dict[rank_list[3]] = 3
            result_dict[rank_list[4]] = 2
            result_dict[rank_list[5]] = 2
            result_dict[rank_list[6]] = 1
            result_dict[rank_list[7]] = 1
            result_dict[rank_list[8]] = 1
        else:
            print("length ERROR!")
        print("同一攻击信息的伤亡情况分级评分")
        print(result_dict)
        return result_dict

    # 获取单一攻击信息下的经济损失等级情况
    def get_economic_loss_by_attack(self):
        propscore_list = self.get_propscore_list()
        result_dict = {}
        attack_times_dict = {}  # 计算某种武器类型的次数
        for i in range(1, NUM_TYPE_ATTACK + 1):
            result_dict[i] = 0
            attack_times_dict[i] = 0
        for i in range(2, self.nrow_data + 1):
            j = i - 2
            if propscore_list[j] != 0:
                score = propscore_list[j]
                attack_type1 = self.sheet.cell(i, self.ncol_attacktype1).value
                attack_type2 = self.sheet.cell(i, self.ncol_attacktype2).value
                attack_type3 = self.sheet.cell(i, self.ncol_attacktype3).value
                if attack_type2 == 0:  # 只有一种攻击的情况
                    m = result_dict[attack_type1]
                    result_dict[attack_type1] = m + score
                    attack_times_dict[attack_type1] += 1
                elif attack_type3 == 0:  # 有两种攻击的情况
                    m = result_dict[attack_type1]
                    o = result_dict[attack_type2]
                    result_dict[attack_type1] = m + score * 1 / 2
                    result_dict[attack_type2] = o + score * 1 / 2
                    attack_times_dict[attack_type1] += 1
                    attack_times_dict[attack_type2] += 1
                else:  # 有三种攻击的情况
                    m = result_dict[attack_type1]
                    o = result_dict[attack_type2]
                    x = result_dict[attack_type3]
                    result_dict[attack_type1] = m + score * 1 / 3
                    result_dict[attack_type2] = o + score * 1 / 3
                    result_dict[attack_type3] = x + score * 1 / 3
                    attack_times_dict[attack_type1] += 1
                    attack_times_dict[attack_type2] += 1
                    attack_times_dict[attack_type3] += 1
        # 将攻击总经济损失得分除去总次数，得出平均威力
        for i in range(1, NUM_TYPE_ATTACK + 1):
            times = attack_times_dict[i]
            m = result_dict[i]
            if times != 0:
                result_dict[i] = m / times * 10000
        rank_list = self.sort_by_value(result_dict)
        result_dict = {}
        if len(rank_list) == NUM_TYPE_ATTACK:
            result_dict[rank_list[0]] = 5
            result_dict[rank_list[1]] = 4
            result_dict[rank_list[2]] = 3
            result_dict[rank_list[3]] = 3
            result_dict[rank_list[4]] = 2
            result_dict[rank_list[5]] = 2
            result_dict[rank_list[6]] = 1
            result_dict[rank_list[7]] = 1
            result_dict[rank_list[8]] = 1
        else:
            print("length ERROR!")
        print("同一攻击信息的经济损失情况分级评分")
        print(result_dict)
        return result_dict

    # 获取单一武器信息下的伤亡情况
    def get_casualty_by_weapon(self):
        num_peop_matrix = self.get_matrix_peop_by_year_and_region()
        result_dict = {}
        weapon_times_dict = {}  # 计算某种武器类型的次数
        for i in range(1, NUM_TYPE_WEAPON + 1):
            result_dict[i] = [0, 0]
            weapon_times_dict[i] = 0
        for i in range(2, self.nrow_data + 1):
            iyear = self.sheet.cell(i, self.ncol_iyear).value
            region = self.sheet.cell(i, self.ncol_region).value
            total = num_peop_matrix[iyear, region]
            nkill = self.sheet.cell(i, self.ncol_nkill).value / total
            nwound = self.sheet.cell(i, self.ncol_nwound).value / total
            weapon_type1 = self.sheet.cell(i, self.ncol_weaptype1).value
            weapon_type2 = self.sheet.cell(i, self.ncol_weaptype2).value
            weapon_type3 = self.sheet.cell(i, self.ncol_weaptype3).value
            if weapon_type2 == 0:  # 只有一种攻击的情况
                m = result_dict[weapon_type1][0]
                n = result_dict[weapon_type1][1]
                result_dict[weapon_type1] = [m + nkill, n + nwound]
                weapon_times_dict[weapon_type1] += 1
            elif weapon_type3 == 0:  # 有两种攻击的情况
                m = result_dict[weapon_type1][0]
                n = result_dict[weapon_type1][1]
                o = result_dict[weapon_type2][0]
                p = result_dict[weapon_type2][1]
                result_dict[weapon_type1] = [m + nkill * 1 / 2, n + nwound * 1 / 2]
                result_dict[weapon_type2] = [o + nkill * 1 / 2, p + nwound * 1 / 2]
                weapon_times_dict[weapon_type1] += 1
                weapon_times_dict[weapon_type2] += 1
            else:  # 有三种攻击的情况
                m = result_dict[weapon_type1][0]
                n = result_dict[weapon_type1][1]
                o = result_dict[weapon_type2][0]
                p = result_dict[weapon_type2][1]
                x = result_dict[weapon_type3][0]
                y = result_dict[weapon_type3][1]
                result_dict[weapon_type1] = [m + nkill * 1 / 3, n + nwound * 1 / 3]
                result_dict[weapon_type2] = [o + nkill * 1 / 3, p + nwound * 1 / 3]
                result_dict[weapon_type3] = [x + nkill * 1 / 3, y + nwound * 1 / 3]
                weapon_times_dict[weapon_type1] += 1
                weapon_times_dict[weapon_type2] += 1
                weapon_times_dict[weapon_type3] += 1
        # 将攻击总威力除去总次数，得出平均威力
        for i in range(1, NUM_TYPE_WEAPON + 1):
            times = weapon_times_dict[i]
            m = result_dict[i][0]
            n = result_dict[i][1]
            if times == 0:
                result_dict[i] = [0, 0]
            else:
                result_dict[i] = [m / times * 10000, n / times * 10000]
        return result_dict

    # 根据伤亡情况对武器信息进行分类
    def get_score_weapon_list_by_casualty(self):
        casualty_dict = self.get_casualty_by_weapon()
        score_dict = {}
        for i in range(1, NUM_TYPE_WEAPON + 1):
            score_dict[i] = casualty_dict[i][0] * 0.8 + casualty_dict[i][1] * 0.2
        rank_list = self.sort_by_value(score_dict)
        result_dict = {}
        if len(rank_list) == NUM_TYPE_WEAPON:
            result_dict[rank_list[0]] = 5
            result_dict[rank_list[1]] = 4
            result_dict[rank_list[2]] = 4
            result_dict[rank_list[3]] = 3
            result_dict[rank_list[4]] = 3
            result_dict[rank_list[5]] = 3
            result_dict[rank_list[6]] = 2
            result_dict[rank_list[7]] = 2
            result_dict[rank_list[8]] = 2
            result_dict[rank_list[9]] = 1
            result_dict[rank_list[10]] = 1
            result_dict[rank_list[11]] = 1
            result_dict[rank_list[12]] = 1
        else:
            print("length ERROR!")
        print("同一武器信息的伤亡情况分级评分")
        print(result_dict)
        return result_dict

    # 获取单一攻击信息下的经济损失等级情况
    def get_economic_loss_by_weapon(self):
        propscore_list = self.get_propscore_list()
        result_dict = {}
        weapon_times_dict = {}  # 计算某种武器类型的次数
        for i in range(1, NUM_TYPE_WEAPON + 1):
            result_dict[i] = 0
            weapon_times_dict[i] = 0
        for i in range(2, self.nrow_data + 1):
            j = i - 2
            if propscore_list[j] != 0:
                score = propscore_list[j]
                weapon_type1 = self.sheet.cell(i, self.ncol_weaptype1).value
                weapon_type2 = self.sheet.cell(i, self.ncol_weaptype2).value
                weapon_type3 = self.sheet.cell(i, self.ncol_weaptype3).value
                if weapon_type2 == 0:  # 只有一种攻击的情况
                    m = result_dict[weapon_type1]
                    result_dict[weapon_type1] = m + score
                    weapon_times_dict[weapon_type1] += 1
                elif weapon_type3 == 0:  # 有两种攻击的情况
                    m = result_dict[weapon_type1]
                    o = result_dict[weapon_type2]
                    result_dict[weapon_type1] = m + score * 1 / 2
                    result_dict[weapon_type2] = o + score * 1 / 2
                    weapon_times_dict[weapon_type1] += 1
                    weapon_times_dict[weapon_type2] += 1
                else:  # 有三种攻击的情况
                    m = result_dict[weapon_type1]
                    o = result_dict[weapon_type2]
                    x = result_dict[weapon_type3]
                    result_dict[weapon_type1] = m + score * 1 / 3
                    result_dict[weapon_type2] = o + score * 1 / 3
                    result_dict[weapon_type3] = x + score * 1 / 3
                    weapon_times_dict[weapon_type1] += 1
                    weapon_times_dict[weapon_type2] += 1
                    weapon_times_dict[weapon_type3] += 1
        # 将攻击总经济损失得分除去总次数，得出平均威力
        for i in range(1, NUM_TYPE_WEAPON + 1):
            times = weapon_times_dict[i]
            m = result_dict[i]
            if times != 0:
                result_dict[i] = m / times * 10000
        rank_list = self.sort_by_value(result_dict)
        result_dict = {}
        if len(rank_list) == NUM_TYPE_WEAPON:
            result_dict[rank_list[0]] = 5
            result_dict[rank_list[1]] = 4
            result_dict[rank_list[2]] = 4
            result_dict[rank_list[3]] = 3
            result_dict[rank_list[4]] = 3
            result_dict[rank_list[5]] = 3
            result_dict[rank_list[6]] = 2
            result_dict[rank_list[7]] = 2
            result_dict[rank_list[8]] = 2
            result_dict[rank_list[9]] = 1
            result_dict[rank_list[10]] = 1
            result_dict[rank_list[11]] = 1
            result_dict[rank_list[12]] = 1
        else:
            print("length ERROR!")
        print("同一武器信息的经济损失情况分级评分")
        print(result_dict)
        return result_dict

    # 获取经济换算得分列表
    def get_propscore_list(self):
        propextent_list = self.get_data_list("propextent")
        propscore_list = self.propextent_to_score(propextent_list)
        num_score = 0
        total_score = 0
        for i in propscore_list:
            if i != 4:
                total_score += i
                num_score += 1
        average_score = total_score / num_score
        for i in range(len(propscore_list)):
            if propscore_list[i] == 4:
                propscore_list[i] = average_score
        return propscore_list

    # 获取指定时间下地区总人数
    def get_matrix_peop_by_year_and_region(self):
        nrow = self.reference_sheet.max_row
        first_list = []
        for i in range(1, nrow + 1):
            first_list.append(self.reference_sheet.cell(1, i).value)
        ncol_iyear = first_list.index('iyear') + 1
        ncol_region = first_list.index('region') + 1
        ncol_num = first_list.index('num_people') + 1
        year_list = []
        region_list = []
        num_list = []
        for i in range(2, self.reference_sheet.max_column + 1):
            year_list.append(self.reference_sheet.cell(i, ncol_iyear).value)
            region_list.append(self.reference_sheet.cell(i, ncol_region).value)
            num_list.append(self.reference_sheet.cell(i, ncol_num).value)
        data_matrix = {}
        if len(year_list) == len(region_list) and len(region_list) == len(num_list):
            for i in range(len(year_list)):
                data_matrix[int(year_list[i]), int(region_list[i])] = num_list[i]
            return data_matrix
        else:
            print("get_matrix_peop_by_year_and_region ERROR!")
            return 0

    # 获取攻击类型的权重分类得分
    def get_matrix_attacktype_score(self):
        result = {}
        casualty_dict = self.get_score_attack_list_by_casualty()
        economic_dict = self.get_economic_loss_by_attack()
        for i in range(1, NUM_TYPE_ATTACK + 1):
            result[i] = casualty_dict[i] * 0.7 + economic_dict[i] * 0.3
        print("单一攻击类型的得分")
        print(result)
        return result

    # 获取攻击类型的权重分类得分
    def get_matrix_weapontype_score(self):
        result = {}
        casualty_dict = self.get_score_weapon_list_by_casualty()
        economic_dict = self.get_economic_loss_by_weapon()
        for i in range(1, NUM_TYPE_WEAPON + 1):
            result[i] = casualty_dict[i] * 0.7 + economic_dict[i] * 0.3
        print("单一武器类型的得分")
        print(result)
        return result

    # 返回两个列表之间的相似度
    @staticmethod
    def ret_similarity(list1, list2):
        if len(list1) != len(list2):
            print("Error!!!!!!")
            return 100
        else:
            res = pow((list1[1] - list2[1]), 2) + pow((list1[2] - list2[2]), 2) + 0.7 * pow((list1[3] - list2[3]),
                                                                                            2) + 0.2 * pow(
                (list1[4] - list2[4]), 2) + 0.1 * pow((list1[5] - list2[5]), 2) + 0.7 * pow((list1[6] - list2[6]),
                                                                                            2) + 0.2 * pow(
                (list1[7] - list2[7]), 2) + 0.1 * pow((list1[8] - list2[8]), 2) + 0.7 * pow((list1[9] - list2[9]),
                                                                                            2) + 0.2 * pow(
                (list1[10] - list2[10]), 2) + 0.1 * pow((list1[11] - list2[11]), 2)
            return sqrt(res)

    # 将经济损失等级换算为经济损失得分
    @staticmethod
    def propextent_to_score(extent_list):
        result_list = []
        for i in extent_list:
            if i == 1:
                result_list.append(1000)
            elif i == 2:
                result_list.append(100)
            elif i == 3:
                result_list.append(10)
            else:
                result_list.append(i)
        return result_list

    # 获取攻击类型的权重分类得分
    @staticmethod
    def get_matrix_attacktype_score_result():
        result = {1: 1.6, 2: 2.7, 3: 2.3, 4: 2.2, 5: 2.4, 6: 1.7, 7: 1.3, 8: 3.8, 9: 4.0}
        print("单一攻击类型的得分")
        print(result)
        return result

    # 获取武器类型的权重分类得分
    @staticmethod
    def get_matrix_weapontype_score_result():
        result = {1: 1.0, 2: 2.3, 3: 1.0, 4: 1.0, 5: 3.7, 6: 3.3, 7: 2.4, 8: 2.3, 9: 2.7, 10: 4.1, 11: 2.2, 12: 2.0,
                  13: 4.0}
        print("单一武器类型的得分")
        print(result)
        return result

    # 获取目标类型的权重分类得分
    @staticmethod
    def get_matrix_targtype_score_result():
        result = {1: 3, 2: 5, 3: 5, 4: 5, 5: 2, 6: 5, 7: 5, 8: 3, 9: 4, 10: 2, 11: 4, 12: 4, 13: 2, 14: 2, 15: 4, 16: 4,
                  17: 3, 18: 2, 19: 4, 20: 1, 21: 4, 22: 3}
        print("单一目标类型的得分")
        print(result)
        return result

    # 获取不同地区的权重分类得分
    @staticmethod
    def get_matrix_region_score_result():
        result = {1: 5, 2: 2, 3: 3, 4: 4, 5: 2, 6: 1, 7: 2, 8: 5, 9: 3, 10: 3, 11: 1, 12: 4}
        print("不同地区的得分")
        print(result)
        return result

    # 对字典的values进行从大到小排序，返回keys
    @staticmethod
    def sort_by_value(d):
        items = d.items()
        backitems = [[v[1], v[0]] for v in items]
        backitems.sort(reverse=True)
        return [backitems[i][1] for i in range(0, len(backitems))]

    # 根据正交因子模型计算
    def get_orthogonal_factor_matrix(self):
        data_dict = self.get_dict_factors_score()
        X = pd.DataFrame(data_dict)
        X_mean = X.mean()
        print("X_mean")
        print(X_mean)
        # 样本相关阵R
        R = X.corr()
        # 求特征值和特征向量
        eig_value, eig_vector = nlg.eig(R)
        eig = pd.DataFrame()
        eig['names'] = X.columns
        eig['eig_value'] = eig_value
        eig.sort_values('eig_value', ascending=False, inplace=True)
        # 求因子模型的因子载荷阵A
        num_factor = 0
        for m in range(1, 11):
            if eig['eig_value'][:m].sum() / eig['eig_value'].sum() > 0.8:
                num_factor = m
                print("num_factor: " + str(m))
                break
        A = np.mat(np.zeros((10, num_factor)))
        for i in range(num_factor):
            A[:, i] = (sqrt(eig_value[i]) * eig_vector[:, i]).reshape(10, 1)
        # 求特殊因子方差和共同度的估计
        h = np.zeros(10)  # 共同度
        D = np.mat(np.eye(10))  # 特殊因子方差
        for i in range(10):
            a = A[i, :] * A[i, :].T
            h[i] = a[0, 0]
            D[i, i] = 1 - a[0, 0]
        a = pd.DataFrame(A)
        column_list = []
        index_list = []
        for i in range(1, num_factor + 1):
            column_list.append('factor' + str(i))
        for j in range(1, 11):
            index_list.append('x' + str(j))
        a.columns = column_list
        a.index = index_list
        print(a)
        # 计算逆矩阵，得出Factor = B * X
        B = A.I
        b = pd.DataFrame(B)
        b.columns = index_list
        b.index = column_list
        print(b)
        # 得出特殊因子得分
        score_result = {}
        for i in range(0, num_factor):
            score = 0
            for j in range(0, 10):
                score += B[i, j] * X_mean[j]
            score_result[i] = score
        print("因子得分列表")
        print(score_result)
        return score_result

    # 获取10个影响因素根据源数据换算分数的dict
    def get_dict_factors_score(self):
        attack_dict = self.get_matrix_attacktype_score_result()
        weapon_dict = self.get_matrix_weapontype_score_result()
        target_dict = self.get_matrix_targtype_score_result()
        region_dict = self.get_matrix_region_score_result()
        attack1 = self.get_data_list("attacktype1")
        for i in range(len(attack1)):
            if attack1[i] != 0:
                attack1[i] = attack_dict[attack1[i]]
        attack2 = self.get_data_list("attacktype2")
        for i in range(len(attack2)):
            if attack2[i] != 0:
                attack2[i] = attack_dict[attack2[i]]
        attack3 = self.get_data_list("attacktype3")
        for i in range(len(attack3)):
            if attack3[i] != 0:
                attack3[i] = attack_dict[attack3[i]]
        weapon1 = self.get_data_list('weaptype1')
        for i in range(len(weapon1)):
            if weapon1[i] != 0:
                weapon1[i] = weapon_dict[weapon1[i]]
        weapon2 = self.get_data_list('weaptype2')
        for i in range(len(weapon2)):
            if weapon2[i] != 0:
                weapon2[i] = weapon_dict[weapon2[i]]
        weapon3 = self.get_data_list('weaptype3')
        for i in range(len(weapon3)):
            if weapon3[i] != 0:
                weapon3[i] = weapon_dict[weapon3[i]]
        target1 = self.get_data_list('targtype1')
        for i in range(len(target1)):
            if target1[i] != 0:
                target1[i] = target_dict[target1[i]]
        target2 = self.get_data_list('targtype2')
        for i in range(len(target2)):
            if target2[i] != 0:
                target2[i] = target_dict[target2[i]]
        target3 = self.get_data_list('targtype3')
        for i in range(len(target3)):
            if target3[i] != 0:
                target3[i] = target_dict[target3[i]]
        region = self.get_data_list('region')
        for i in range(len(region)):
            if region[i] != 0:
                region[i] = region_dict[region[i]]
        data_dict = {'attack1': attack1, 'attack2': attack2, 'attack3': attack3, 'weapon1': weapon1, 'weapon2': weapon2,
                     'weapon3': weapon3, 'target1': target1, 'target2': target2, 'target3': target3,
                     'region': region}
        return data_dict

    # 获取恐怖事件的危害程度得分
    def get_list_terror_score(self):
        data_dict = self.get_dict_factors_score()


